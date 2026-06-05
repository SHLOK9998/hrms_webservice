from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, date
from typing import Optional
from utils.timezone import get_current_time, get_current_date
from bson import ObjectId
from database import get_database
from schemas import AttendanceCreate
from utils.auth import get_current_admin, get_current_user
import io

router = APIRouter(prefix="/api/attendance", tags=["Attendance"])

def serialize(doc):
    doc["_id"] = str(doc["_id"])
    return doc

@router.get("/")
async def get_all_attendance(
    employee_id: Optional[str] = Query(None),
    month: Optional[str] = Query(None),  # format: "YYYY-MM"
    current_user=Depends(get_current_admin),
):
    db = get_database()
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if month:
        query["date"] = {"$regex": f"^{month}"}
    records = await db.attendance.find(query).sort("date", -1).to_list(2000)
    return [serialize(r) for r in records]

@router.get("/my")
async def get_my_attendance(current_user=Depends(get_current_user)):
    db = get_database()
    employee = await db.employees.find_one({"email": current_user["email"]})
    if not employee:
        return []
    records = await db.attendance.find({"employee_id": employee["employee_id"]}).sort("date", -1).to_list(100)
    return [serialize(r) for r in records]

@router.get("/today")
async def get_today_attendance(current_user=Depends(get_current_admin)):
    db = get_database()
    today = get_current_date().isoformat()
    records = await db.attendance.find({"date": today}).to_list(1000)
    return [serialize(r) for r in records]

@router.get("/today/count")
async def get_today_checkin_count(current_user=Depends(get_current_admin)):
    """Return count of employees who checked in today."""
    db = get_database()
    today = get_current_date().isoformat()
    count = await db.attendance.count_documents({"date": today, "status": "present"})
    return {"checked_in_today": count}

@router.get("/today/status")
async def get_my_today_status(current_user=Depends(get_current_user)):
    """Return the current employee's check-in status for today."""
    db = get_database()
    employee = await db.employees.find_one({"email": current_user["email"]})
    if not employee:
        return {"checked_in": False, "checked_out": False, "record": None}
    today = get_current_date().isoformat()
    record = await db.attendance.find_one({"employee_id": employee["employee_id"], "date": today})
    if not record:
        return {"checked_in": False, "checked_out": False, "record": None}
    record = serialize(record)
    return {
        "checked_in": True,
        "checked_out": bool(record.get("check_out")),
        "record": record,
    }

@router.post("/checkin")
async def check_in(current_user=Depends(get_current_user)):
    db = get_database()
    employee = await db.employees.find_one({"email": current_user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee profile not found")
    today = get_current_date().isoformat()

    # ── Block check-in on approved leave day ──────────────────────────────
    on_leave = await db.leaves.find_one({
        "employee_email": current_user["email"],
        "status": "approved",
        "start_date": {"$lte": today},
        "end_date": {"$gte": today},
    })
    if on_leave:
        raise HTTPException(
            status_code=400,
            detail="You cannot check in during an approved leave day"
        )

    existing = await db.attendance.find_one({"employee_id": employee["employee_id"], "date": today})
    if existing:
        raise HTTPException(status_code=400, detail="Already checked in today")
    now = get_current_time()
    record = {
        "employee_id": employee["employee_id"],
        "employee_name": employee["full_name"],
        "employee_email": current_user["email"],
        "department": employee.get("department", ""),
        "date": today,
        "check_in": now.strftime("%H:%M:%S"),
        "check_out": None,
        "total_hours": None,
        "status": "present",
        "created_at": now,
    }
    result = await db.attendance.insert_one(record)
    record["_id"] = str(result.inserted_id)
    return record

@router.post("/checkout")
async def check_out(current_user=Depends(get_current_user)):
    db = get_database()
    employee = await db.employees.find_one({"email": current_user["email"]})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee profile not found")
    today = get_current_date().isoformat()
    existing = await db.attendance.find_one({"employee_id": employee["employee_id"], "date": today})
    if not existing:
        raise HTTPException(status_code=400, detail="No check-in found for today")
    if existing.get("check_out"):
        raise HTTPException(status_code=400, detail="Already checked out today")

    now = get_current_time()
    # Calculate total hours
    check_in_time = datetime.strptime(existing["check_in"], "%H:%M:%S")
    check_out_time = datetime.strptime(now.strftime("%H:%M:%S"), "%H:%M:%S")
    delta = check_out_time - check_in_time
    total_hours = round(delta.total_seconds() / 3600, 2)

    await db.attendance.update_one(
        {"_id": existing["_id"]},
        {"$set": {"check_out": now.strftime("%H:%M:%S"), "total_hours": total_hours, "updated_at": now}}
    )
    return {"message": "Checked out successfully", "check_out": now.strftime("%H:%M:%S"), "total_hours": total_hours}

@router.post("/")
async def mark_attendance(data: AttendanceCreate, current_user=Depends(get_current_admin)):
    db = get_database()
    existing = await db.attendance.find_one({"employee_id": data.employee_id, "date": data.date})
    if existing:
        await db.attendance.update_one({"_id": existing["_id"]}, {"$set": data.model_dump()})
        return {"message": "Attendance updated"}
    employee = await db.employees.find_one({"employee_id": data.employee_id})
    record = {
        **data.model_dump(),
        "employee_name": employee["full_name"] if employee else "Unknown",
        "employee_email": employee["email"] if employee else "",
        "department": employee.get("department", "") if employee else "",
        "created_at": get_current_time(),
    }
    result = await db.attendance.insert_one(record)
    record["_id"] = str(result.inserted_id)
    return record

@router.get("/stats/monthly")
async def get_monthly_stats(current_user=Depends(get_current_admin)):
    db = get_database()
    current_month = get_current_date().strftime("%Y-%m")
    pipeline = [
        {"$match": {"date": {"$regex": f"^{current_month}"}}},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
    ]
    result = await db.attendance.aggregate(pipeline).to_list(10)
    return result

@router.get("/stats/avg-hours")
async def get_avg_hours_stats(current_user=Depends(get_current_admin)):
    """Average hours per day for the current month."""
    db = get_database()
    current_month = get_current_date().strftime("%Y-%m")
    pipeline = [
        {"$match": {"date": {"$regex": f"^{current_month}"}, "total_hours": {"$ne": None}}},
        {"$group": {
            "_id": "$department",
            "avg_hours": {"$avg": "$total_hours"},
            "total_days": {"$sum": 1}
        }},
        {"$sort": {"avg_hours": -1}}
    ]
    result = await db.attendance.aggregate(pipeline).to_list(20)
    return result


# ── Excel Report Download ─────────────────────────────────────────────────────

@router.get("/export/excel")
async def export_attendance_excel(
    employee_id: Optional[str] = Query(None),
    month: Optional[str] = Query(None),
    current_user=Depends(get_current_admin),
):
    """Download attendance records as an Excel (.xlsx) file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed. Run: pip install openpyxl")

    db = get_database()
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if month:
        query["date"] = {"$regex": f"^{month}"}

    records = await db.attendance.find(query).sort("date", -1).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Employee ID", "Employee Name", "Department", "Date", "Check In", "Check Out", "Total Hours", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, r in enumerate(records, 2):
        values = [
            r.get("employee_id", ""),
            r.get("employee_name", ""),
            r.get("department", ""),
            r.get("date", ""),
            r.get("check_in", ""),
            r.get("check_out", ""),
            r.get("total_hours", ""),
            r.get("status", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Auto-width
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    emp_label = employee_id or "all"
    month_label = month or "all"
    filename = f"attendance_{emp_label}_{month_label}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
